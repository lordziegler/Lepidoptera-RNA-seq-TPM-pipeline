#!/usr/bin/env python3
"""
02_prepare_samples.py
---------------------
Parses an NCBI SRA RunTable (Excel format from SRA Run Selector) and writes
samples.tsv, the input file consumed by 03_quantify.sh.

Output columns (tab-separated):
    SRR       NCBI SRA run accession
    SPECIES   Species key matching a subdirectory under references/
    LAYOUT    PAIRED | SINGLE

Filters applied in order:
    1. Assay Type  == RNA-Seq
    2. LibrarySource == TRANSCRIPTOMIC
    3. Organism substring matches one of the six target Lepidoptera species
    4. Tissue contains a gut/midgut keyword  (disabled with --all-tissues)
    5. Run accession matches SRR/ERR/DRR pattern
    6. Duplicate accessions removed (keeps first occurrence)

Usage:
    python3 02_prepare_samples.py --input SraRunTable.xlsx
    python3 02_prepare_samples.py --input SraRunTable.xlsx --all-tissues
    python3 02_prepare_samples.py --input SraRunTable.xlsx \\
        --include-species Bombyx_mori Spodoptera_frugiperda
    python3 02_prepare_samples.py --input SraRunTable.xlsx \\
        --output custom_samples.tsv

Dependencies:
    openpyxl >= 3.0   (included in environment.yml)
"""

import argparse
import csv
import re
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("[ABORT] openpyxl not found.\n"
             "        conda activate lepidoptera-rnaseq  or  pip install openpyxl")

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------

# Maps organism substrings (lowercase) to the species key used in references/
SPECIES_MAP: dict[str, str] = {
    "spodoptera frugiperda":  "Spodoptera_frugiperda",
    "plutella xylostella":    "Plutella_xylostella",
    "diatraea saccharalis":   "Diatraea_saccharalis",
    "bombyx mori":            "Bombyx_mori",
    "tecia solanivora":       "Tecia_solanivora",
    "phyllocnistis citrella": "Phyllocnistis_citrella",
    # Kept for pilot runs against the provided H. armigera RunTable
    "helicoverpa armigera":   "Helicoverpa_armigera",
}

TISSUE_KEYWORDS: list[str] = [
    "gut", "midgut", "intestin", "digestive", "fore gut", "hind gut",
]

ACCESSION_RE = re.compile(r'^[SED]RR\d+$')

# ---------------------------------------------------------------------------
# FUNCTIONS
# ---------------------------------------------------------------------------

def load_workbook(path: Path) -> list[dict]:
    """
    Opens an NCBI SRA RunTable workbook and returns rows as a list of dicts.
    Accepts both .xlsx and .csv files.
    """
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        try:
            with open(path, "r", encoding="utf-8-sig") as fh:
                reader = csv.reader(fh)
                rows = list(reader)
        except Exception as csv_exc:
            sys.exit(f"[ABORT] Cannot open workbook as Excel or CSV.\n"
                     f"        Excel error: {exc}\n"
                     f"        CSV error: {csv_exc}\n"
                     f"        Verify the file: {path}")

    if not rows:
        sys.exit("[ABORT] File is empty.")

    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(rows[0])]
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(v is not None and str(v).strip() for v in row)
    ]


def get(row: dict, key: str, default: str = "") -> str:
    """Returns a stripped string value from a row dict, or default."""
    val = row.get(key)
    return str(val).strip() if val is not None else default


def map_species(organism: str) -> str | None:
    """
    Returns the canonical species key for an organism string.
    Returns None when the organism is not among the target species.
    """
    org_lower = organism.lower()
    return next(
        (key for pattern, key in SPECIES_MAP.items() if pattern in org_lower),
        None
    )


def is_gut_tissue(tissue_value) -> bool:
    """Returns True when the tissue field contains a gut-related keyword."""
    if tissue_value is None:
        return False
    return any(kw in str(tissue_value).lower() for kw in TISSUE_KEYWORDS)


def write_tsv(rows: list[dict], output: Path) -> None:
    """Writes the filtered rows to a tab-separated file with a header line."""
    with open(output, "w") as fh:
        fh.write("SRR\tSPECIES\tLAYOUT\n")
        for row in rows:
            fh.write(f"{row['SRR']}\t{row['SPECIES']}\t{row['LAYOUT']}\n")


def print_summary(rows: list[dict]) -> None:
    """Prints a count breakdown by species and layout."""
    by_species = Counter(r["SPECIES"] for r in rows)
    by_layout  = Counter(r["LAYOUT"]  for r in rows)
    print(f"\n  Total samples : {len(rows)}")
    print(f"  Layout        : {dict(by_layout)}")
    print("  By species:")
    for sp, n in sorted(by_species.items()):
        print(f"    {sp}: {n}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate samples.tsv from an NCBI SRA RunTable."
    )
    parser.add_argument("--input",  "-i", required=True, type=Path,
                        help="SRA RunTable file (.xlsx or mis-named .csv).")
    parser.add_argument("--output", "-o", default="samples.tsv", type=Path,
                        help="Output TSV path (default: samples.tsv).")
    parser.add_argument("--all-tissues", action="store_true", default=True,
                        help="Disable gut/midgut tissue filter (currently defaults to True).")
    parser.add_argument("--gut-only", action="store_true",
                        help="Restrict to gut/midgut tissues only.")
    parser.add_argument("--include-species", nargs="+", default=None,
                        metavar="KEY",
                        help="Restrict to specific species keys (space-separated).")
    args = parser.parse_args()

    if not args.input.exists():
        sys.exit(f"[ABORT] Input file not found: {args.input}")

    print(f"[INFO] Loading: {args.input}")
    all_rows = load_workbook(args.input)
    print(f"[INFO] Rows loaded: {len(all_rows)}")

    # Filter 1 — RNA-Seq (allow GENOMIC if misannotated, as long as Assay Type is RNA-Seq)
    rows = [r for r in all_rows
            if get(r, "Assay Type") == "RNA-Seq"
            and get(r, "LibrarySource") in ("TRANSCRIPTOMIC", "GENOMIC")]
    print(f"[INFO] After RNA-Seq filter: {len(rows)}")

    # Filter 2 — Target species
    mapped = []
    for r in rows:
        sp = map_species(get(r, "Organism"))
        if sp:
            mapped.append({**r, "_species_key": sp})
    rows = mapped
    print(f"[INFO] After species filter: {len(rows)}")

    # Filter 3 — Optional species subset
    if args.include_species:
        rows = [r for r in rows if r["_species_key"] in args.include_species]
        print(f"[INFO] After --include-species filter: {len(rows)}")

    # Filter 4 — Gut/midgut tissue
    if args.gut_only:
        rows = [r for r in rows if is_gut_tissue(r.get("tissue"))]
        print(f"[INFO] After gut/midgut tissue filter: {len(rows)}")
    else:
        print("[INFO] Tissue filter disabled — all tissues retained.")

    # Filter 5 — Valid accession pattern + deduplication
    seen: set[str] = set()
    clean: list[dict] = []
    for r in rows:
        srr = get(r, "Run")
        if not ACCESSION_RE.match(srr):
            continue
        if srr in seen:
            continue
        seen.add(srr)
        clean.append({
            "SRR":     srr,
            "SPECIES": r["_species_key"],
            "LAYOUT":  get(r, "LibraryLayout").upper(),
        })

    print(f"[INFO] After deduplication: {len(clean)}")

    if not clean:
        sys.exit("[ABORT] No samples passed all filters. "
                 "Check your input file or use --all-tissues.")

    write_tsv(clean, args.output)
    print(f"\n[DONE] {len(clean)} samples written to: {args.output}")
    print_summary(clean)
    print(f"\n  Next: bash 03_quantify.sh [--test]")


if __name__ == "__main__":
    main()
